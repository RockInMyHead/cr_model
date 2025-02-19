from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.contrib.auth.decorators import login_required
from .forms import RegistrationForm, Predict, ReportsForm
from .models import Transactions, Services, Reports, Notification
from django.core.paginator import Paginator
from django.utils.dateparse import parse_date
from sklearn.preprocessing import LabelEncoder
from tensorflow.keras.models import load_model
import openpyxl
from openpyxl.styles import Font
import os
import numpy as np
import pandas as pd
from django.conf import settings
from django.shortcuts import render
from django.contrib.admin.models import LogEntry

from datetime import datetime

# Load the trained model
MODEL_PATH = os.path.join(settings.BASE_DIR, 'users', 'best_model_3.keras')
model = load_model(MODEL_PATH)


def event_history_view(request):
    # Получаем последние 20 записей лога действий, отсортированные по времени действия (самые свежие – первыми)
    events = LogEntry.objects.all().order_by('-action_time')[:20]
    return render(request, 'users/update.html', {'events': events})

def register(request):
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            form.save()
            # (Опционально: можно добавить уведомление о регистрации)
            return redirect('users:login')
    else:
        form = RegistrationForm()
    return render(request, 'users/register.html', {'form': form})


def log_notification(user, text):
    """
    Функция для создания уведомления для пользователя.
    """
    if user.is_authenticated:
        Notification.objects.create(user=user, text=text)


def login(request):
    return render(request, 'users/login.html')


def index(request):
    return render(request, 'users/index.html')


@login_required
def test(request):
    # 1. Обработка POST-запроса для добавления отчёта
    if request.method == 'POST' and 'add_report' in request.POST:
        form = ReportsForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            report.save()
            log_notification(request.user, "Вы добавили новый отчёт.")
            messages.success(request, "Отчёт успешно добавлен!")
            return redirect('users:test')
        else:
            messages.error(request, "Произошла ошибка при сохранении отчёта.")
    else:
        form = ReportsForm()

    # 2. Фильтрация отчётов
    user_reports = Reports.objects.filter(user=request.user)
    report_name = request.GET.get('report_name', '').strip()
    report_date = request.GET.get('report_date', '').strip()
    if report_name:
        user_reports = user_reports.filter(name__icontains=report_name)
    if report_date:
        user_reports = user_reports.filter(date=report_date)

    # 3. Фильтрация транзакций
    user_transactions = Transactions.objects.all()
    transaction_bin = request.GET.get('transaction_bin', '').strip()
    transaction_day = request.GET.get('transaction_day', '').strip()
    transaction_month = request.GET.get('transaction_month', '').strip()
    if transaction_bin:
        user_transactions = user_transactions.filter(bin__icontains=transaction_bin)
    if transaction_day:
        user_transactions = user_transactions.filter(Day=transaction_day)
    if transaction_month:
        user_transactions = user_transactions.filter(Month=transaction_month)

    # 4. Проверка роли пользователя. Если роль установлена и ее имя равно "one",
    # то устанавливаем флаг show_elements в True
    show_elements = False
    if request.user.role:
        show_elements = (request.user.role.name == "one")

    context = {
        'form': form,
        'user_reports': user_reports,
        'user_transactions': user_transactions,
        'show_elements': show_elements,  # передаем флаг в шаблон
    }

    return render(request, 'users/test.html', context)


def about(request):
    return render(request, 'users/about.html')


def get_time_of_day(seconds):
    if 5 * 3600 <= seconds < 12 * 3600:
        return 'morning'
    elif 12 * 3600 <= seconds < 17 * 3600:
        return 'evening'
    else:
        return 'night'


def process_creationdate(df):
    df['creationdate'] = pd.to_datetime(df['creationdate'])
    df['day'] = df['creationdate'].dt.day
    df['month'] = df['creationdate'].dt.month
    return df


def prepare_input(data):
    df = pd.DataFrame([data])

    # Drop unnecessary columns if present
    df = df.drop(columns=['txid', 'card_id', 'currencycode'], errors='ignore')

    # Map 'cardverificationcodesupplied' from 'yes'/'no' to 1/0
    df['cardverificationcodesupplied'] = df['cardverificationcodesupplied'].map({'yes': 1, 'no': 0}).fillna(1).astype(int)

    # Create 'is_foreign_transaction'
    df['is_foreign_transaction'] = (df['issuercountrycode'] != df['shoppercountrycode']).astype(int)

    # Process 'creationdate' if present
    if 'creationdate' in df.columns:
        df = process_creationdate(df)

    # Calculate 'bin_length'
    df['bin_length'] = df['bin'].astype(str).apply(len)
    print("!!!!!!!!!!!!!!!!!!!!!")
    print(df['bin_length'].values)
    # Calculate 'amount_to_country_avg'
    country_avg_amount = df.groupby('shoppercountrycode')['amount'].transform('mean')
    df['amount_to_country_avg'] = np.round(df['amount'] / country_avg_amount, 2).fillna(0)

    # Calculate 'amount_to_card_type_avg'
    card_type_avg_amount = df.groupby('txvariantcode')['amount'].transform('mean')
    df['amount_to_card_type_avg'] = np.round(df['amount'] / card_type_avg_amount, 2).fillna(0)

    # Calculate 'cvcres_avg_amount'
    cvcres_avg_amount = df.groupby('cvcresponsecode')['amount'].transform('mean')
    df['cvcres_avg_amount'] = cvcres_avg_amount.fillna(0)

    # Calculate 'cardver_avg_amount'
    cardver_avg_amount = df.groupby('cardverificationcodesupplied')['amount'].transform('mean')
    df['cardver_avg_amount'] = cardver_avg_amount.fillna(0)

    # Calculate 'is_large_transaction'
    df['is_large_transaction'] = (df['amount'] > 5000).astype(int)

    # Calculate 'is_small_transaction'
    df['is_small_transaction'] = (((df['amount'] > 500) & (df['amount'] < 1500)) |
                                  ((df['amount'] > 3000) & (df['amount'] < 4000))).astype(int)

    # Calculate 'is_holiday_season'
    df['is_holiday_season'] = df['month'].isin([1, 12]).astype(int)

    # Determine 'time_of_day' and create dummies
    df['time_of_day'] = df['time_in_seconds'].apply(get_time_of_day)
    df = pd.get_dummies(df, columns=['time_of_day'], drop_first=True)

    # Calculate 'bin_ver_percentaget'
    bin_ver_percentage = df.groupby('bin')['cardverificationcodesupplied'].transform('mean')
    df['bin_ver_percentaget'] = np.round(bin_ver_percentage, 2).fillna(0)

    # Calculate 'bin_cvs_res_code'
    bin_cvs_res_code = df.groupby('cvcresponsecode')['bin'].transform('mean')
    df['bin_cvs_res_code'] = np.round(bin_cvs_res_code, 2).fillna(0)

    # Calculate 'bin_avg_amount'
    bin_avg_amount = df.groupby('bin')['amount'].transform('mean')
    df['bin_avg_amount'] = np.round(bin_avg_amount, 2).fillna(0)

    # One-hot encode 'shoppercountrycode', 'txvariantcode', 'issuercountrycode'
    shopper_dummies = pd.get_dummies(df['shoppercountrycode'], prefix='shoppercountrycode')
    txvariant_dummies = pd.get_dummies(df['txvariantcode'], prefix='txvariantcode')
    issuercountry_dummies = pd.get_dummies(df['issuercountrycode'], prefix='issuercountrycode')

    # Concatenate dummies with the main dataframe
    df = pd.concat([df, shopper_dummies, txvariant_dummies, issuercountry_dummies], axis=1)

    # Drop original categorical columns
    df = df.drop(columns=['shoppercountrycode', 'txvariantcode', 'issuercountrycode'], errors='ignore')

    # Define all required features (84 features excluding 'is_fraud')
    required_features = [
        'bin', 'amount', 'day', 'month', 'time_in_seconds', 'bin_length',
        'amount_to_country_avg', 'amount_to_card_type_avg', 'bin_avg_amount',
        'bin_ver_percentaget', 'bin_cvs_res_code', 'cvcres_avg_amount',
        'cardver_avg_amount', 'cardverificationcodesupplied',
        'cvcresponsecode', 'is_foreign_transaction', 'is_large_transaction',
        'is_small_transaction', 'is_holiday_season', 'time_of_day_evening',
        'time_of_day_morning', 'time_of_day_night',
        # Issuer country codes
        'issuercountrycode_AR', 'issuercountrycode_AT', 'issuercountrycode_AU',
        'issuercountrycode_BR', 'issuercountrycode_CA', 'issuercountrycode_CL',
        'issuercountrycode_CO', 'issuercountrycode_DE', 'issuercountrycode_EG',
        'issuercountrycode_ES', 'issuercountrycode_ET', 'issuercountrycode_FR',
        'issuercountrycode_GB', 'issuercountrycode_GR', 'issuercountrycode_HR',
        'issuercountrycode_IT', 'issuercountrycode_MA', 'issuercountrycode_MX',
        'issuercountrycode_NI', 'issuercountrycode_NL', 'issuercountrycode_NO',
        'issuercountrycode_NZ', 'issuercountrycode_PL', 'issuercountrycode_PT',
        'issuercountrycode_RO', 'issuercountrycode_SE', 'issuercountrycode_SO',
        'issuercountrycode_UA', 'issuercountrycode_US', 'issuercountrycode_VE',
        # Transaction variant codes
        'txvariantcode_MCDEBIT', 'txvariantcode_VISA',
        'txvariantcode_VISABUSINESS', 'txvariantcode_VISACLASSIC',
        'txvariantcode_VISADEBIT',
        # Shopper country codes
        'shoppercountrycode_AT', 'shoppercountrycode_AU', 'shoppercountrycode_BR',
        'shoppercountrycode_CA', 'shoppercountrycode_CL', 'shoppercountrycode_CO',
        'shoppercountrycode_DE', 'shoppercountrycode_EG', 'shoppercountrycode_ES',
        'shoppercountrycode_ET', 'shoppercountrycode_FR', 'shoppercountrycode_GB',
        'shoppercountrycode_GR', 'shoppercountrycode_HR', 'shoppercountrycode_IT',
        'shoppercountrycode_MA', 'shoppercountrycode_MX', 'shoppercountrycode_NI',
        'shoppercountrycode_NL', 'shoppercountrycode_NZ', 'shoppercountrycode_PT',
        'shoppercountrycode_RO', 'shoppercountrycode_SE', 'shoppercountrycode_SO',
        'shoppercountrycode_UA', 'shoppercountrycode_US', 'shoppercountrycode_VE'
    ]

    # Ensure all required features are present
    for feature in required_features:
        if feature not in df.columns:
            df[feature] = 0
    df = df.drop(columns=['shoppercountrycode_UA', 'shoppercountrycode_US'], errors='ignore')
    # Reorder columns to match the required feature order
    # df = df[required_features]

    # Convert all features to float32
    df = df.astype(float)

    # Debugging: Print the prepared DataFrame
    print("Prepared Input DataFrame:")
    print(df.head())
    print(f"Input Data Shape: {df.shape}")

    return df.values


def predict_transactions(request):
    if request.method == 'POST':
        form = Predict(request.POST)
        if form.is_valid():
            try:
                # Сначала сохраняем введённые пользователем данные в таблицу Transactions
                transaction = Transactions.objects.create(
                    bin=str(form.cleaned_data.get('bin', '0')),
                    amount=str(form.cleaned_data.get('amount', '0')),
                    shoppercountrycode=form.cleaned_data.get('shoppercountrycode', ''),
                    cardverificationcodesupplied=form.cleaned_data.get('cardverificationcodesupplied', 'yes'),
                    cvcresponsecode=str(form.cleaned_data.get('cvcresponsecode', '0')),
                    txvariantcode=form.cleaned_data.get('txvariantcode', ''),
                    Day=str(form.cleaned_data.get('Day', '0')),
                    Month=str(form.cleaned_data.get('Month', '0')),
                    time_in_seconds=str(form.cleaned_data.get('time_in_seconds', '0')),
                    issuercountrycode=form.cleaned_data.get('issuercountrycode', '')
                )

                # Подготовка данных для прогноза (приведение к числовым типам)
                data = {
                    'bin': int(form.cleaned_data.get('bin', 0)),
                    'amount': float(form.cleaned_data.get('amount', 0)),
                    'shoppercountrycode': form.cleaned_data.get('shoppercountrycode', ''),
                    'cardverificationcodesupplied': form.cleaned_data.get('cardverificationcodesupplied', 'yes'),
                    'cvcresponsecode': int(form.cleaned_data.get('cvcresponsecode', 0)),
                    'txvariantcode': form.cleaned_data.get('txvariantcode', ''),
                    'day': int(form.cleaned_data.get('Day', 0)),
                    'month': int(form.cleaned_data.get('Month', 0)),
                    'time_in_seconds': float(form.cleaned_data.get('time_in_seconds', 0)),
                    'issuercountrycode': form.cleaned_data.get('issuercountrycode', '')
                }

                # Debug: вывод исходных данных
                print("Raw Input Data from Form:")
                print(data)

                input_data = prepare_input(data)

                # Debug: вывод подготовленных данных
                print("Input Data Passed to Model:")
                print(input_data)
                print(f"Input Data Shape: {input_data.shape}")

                prediction = model.predict(input_data)
                predicted_class = prediction[0][0]

                # Записываем уведомление о получении данных (если пользователь авторизован)
                if request.user.is_authenticated:
                    log_notification(request.user, "Вы получили данные по транзакции.")

                # Debug: вывод результата предсказания
                print(f"Model Prediction: {predicted_class}")

                return JsonResponse({'prediction': float(predicted_class)})

            except Exception as e:
                print(f"Prediction Error: {e}")
                return JsonResponse({'error': str(e)}, status=400)
        else:
            print("Form is invalid:")
            print(form.errors)
            return JsonResponse({'error': 'Invalid form data'}, status=400)
    else:
        form = Predict()
    return render(request, 'users/predict.html', {'form': form})


def services(request):
    services = Services.objects.all()
    
    # Filter by name
    name = request.GET.get('name', '').strip()
    if name:
        services = services.filter(name__icontains=name)
    
    # Filter by description
    description = request.GET.get('desc', '').strip()
    if description:
        services = services.filter(desc__icontains=description)
    
    # Filter by dates
    start_date = request.GET.get('start_date', '').strip()
    end_date = request.GET.get('end_date', '').strip()
    
    if start_date:
        try:
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
            services = services.filter(updated_at__date__gte=start_date_obj)
        except ValueError:
            messages.error(request, "Неверный формат даты начала. Используйте формат YYYY-MM-DD.")
    
    if end_date:
        try:
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
            services = services.filter(updated_at__date__lte=end_date_obj)
        except ValueError:
            messages.error(request, "Неверный формат даты окончания. Используйте формат YYYY-MM-DD.")
    
    # Pagination
    paginator = Paginator(services, 10)  # 10 services per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'users/serv.html', {'page_obj': page_obj})


# Допустим, этот код в users/utils.py

import numpy as np
import pandas as pd
from tensorflow.keras.models import load_model
from .models import Transactions

def generate(
    country=None,
    card_type=None,
    cvv=None,
    threshold=0.5,
    count=5
):
    import numpy as np
    import pandas as pd
    from tensorflow.keras.models import load_model
    from .models import Transactions

    # Загрузка модели
    model_path = r"D:\diplom_django\cr_model-main\diplom_2\myproject\users\fraud_detection_model.h5"
    model = load_model(model_path)
    expected_features = model.input_shape[1]
    print(f"Модель ожидает {expected_features} признаков.")

    num_samples = count

    # Справочники для преобразования кодов в текст
    month_dict = {
        1: 'Январь', 2: 'Февраль', 3: 'Март',
        4: 'Апрель', 5: 'Май', 6: 'Июнь',
        7: 'Июль', 8: 'Август', 9: 'Сентябрь',
        10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    country_dict = {840: 'США', 643: 'Россия', 250: 'Франция', 356: 'Индия', 36: 'Венгрия', 276: 'Германия'}
    tx_variant_dict = {101: 'Visa', 102: 'Mastercard', 103: 'Maestro', 104: 'Mir'}

    # Генерация базовых данных
    numeric_data = pd.DataFrame({
        'bin': np.random.randint(400000, 500000, size=num_samples),
        'shoppercountrycode': np.random.choice(list(country_dict.keys()), size=num_samples),
        'txvariantcode': np.random.choice(list(tx_variant_dict.keys()), size=num_samples),
        'issuercountrycode': np.random.choice(list(country_dict.keys()), size=num_samples),
        'amount': np.random.randint(100, 10001, size=num_samples),
        'Day': np.random.randint(1, 29, size=num_samples),
        'Month': np.random.randint(1, 13, size=num_samples),
        'time_in_seconds': np.random.randint(0, 86400, size=num_samples),
        # Генерируем CVC Supplied как 0 или 1
        'cardverificationcodesupplied': np.random.randint(0, 2, size=num_samples)
    })

    # Генерация CVC Response:
    # Если CVC supplied == 1, генерируем случайное 3-значное число, иначе пустую строку.
    numeric_data['cvcresponsecode'] = numeric_data['cardverificationcodesupplied'].apply(
        lambda x: str(np.random.randint(100, 1000)) if x == 1 else ""
    )

    # Добавляем столбец bin_length
    numeric_data['bin_length'] = numeric_data['bin'].astype(str).apply(len)

    # Преобразование Month и кодов стран/типов в текст
    numeric_data['Month'] = numeric_data['Month'].apply(lambda x: month_dict[x])
    numeric_data['shoppercountrycode'] = numeric_data['shoppercountrycode'].apply(lambda x: country_dict[x])
    numeric_data['issuercountrycode'] = numeric_data['issuercountrycode'].apply(lambda x: country_dict[x])
    numeric_data['txvariantcode'] = numeric_data['txvariantcode'].apply(lambda x: tx_variant_dict[x])

    # Если переданы параметры, фиксируем соответствующие поля
    if country:
        numeric_data['shoppercountrycode'] = country
    if card_type:
        numeric_data['txvariantcode'] = card_type
    if cvv:
        numeric_data['cvcresponsecode'] = cvv

    # Преобразуем значение CVC Supplied: 0 -> "Нет", 1 -> "Да"
    numeric_data['cardverificationcodesupplied'] = numeric_data['cardverificationcodesupplied'].map({0: "Нет", 1: "Да"})

    # Генерация "модели" для демонстрационных целей (здесь используется случайный DataFrame)
    # Этот блок можно изменить в зависимости от реальной логики подготовки данных для модели
    import copy
    df_for_model = pd.DataFrame()
    feature_cols = [
        'bin', 'amount', 'shoppercountrycode', 'cardverificationcodesupplied',
        'cvcresponsecode', 'txvariantcode', 'Day', 'Month', 'time_in_seconds',
        'issuercountrycode', 'bin_length'
    ]
    df_for_model[feature_cols] = np.random.randint(0, 1000, size=(num_samples, len(feature_cols)))

    features_to_multiply = [
        'shoppercountrycode', 'cardverificationcodesupplied', 'cvcresponsecode',
        'txvariantcode', 'Day', 'Month', 'time_in_seconds', 'issuercountrycode'
    ]
    for col in features_to_multiply:
        multiplier = np.random.uniform(1.0, 10.0)
        df_for_model[col] = df_for_model[col] * multiplier

    model_input_np = df_for_model.to_numpy()
    predictions = model.predict(model_input_np)

    if predictions.ndim > 1 and predictions.shape[1] > 1:
        predicted_values = np.argmax(predictions, axis=1)
    else:
        predicted_values = predictions.flatten()

    # Пример случайной подмены части предсказаний
    total = len(predicted_values)
    count_to_replace = np.random.randint(1, total + 1)
    indices_to_replace = np.random.choice(total, size=count_to_replace, replace=False)
    for idx in indices_to_replace:
        predicted_values[idx] = np.random.randint(1, 100)

    def map_prediction(val):
        if val == 0:
            return "Не мошенническая"
        elif val == 1:
            return "Мошенническая"
        else:
            return "Не известно"

    numeric_data['Prediction'] = predicted_values
    numeric_data['Prediction_result'] = numeric_data['Prediction'].apply(map_prediction)

    # Сохранение транзакций в базе данных с указанием порога предсказания
    for idx, row in numeric_data.iterrows():
        Transactions.objects.create(
            bin=str(row['bin']),
            amount=str(row['amount']),
            shoppercountrycode=str(row['shoppercountrycode']),
            cardverificationcodesupplied=str(row['cardverificationcodesupplied']),
            cvcresponsecode=str(row['cvcresponsecode']),
            txvariantcode=str(row['txvariantcode']),
            Day=str(row['Day']),
            Month=str(row['Month']),
            time_in_seconds=str(row['time_in_seconds']),
            issuercountrycode=str(row['issuercountrycode']),
            result=row['Prediction_result'],
            threshold=threshold  # сохраняем значение порога
        )

    predict_str = f"Создано {num_samples} транзакций. Предсказания: {set(numeric_data['Prediction'])}"
    return predict_str, numeric_data



def generated(request):
    """
    Обрабатывает POST-запросы для:
      - Динамической генерации транзакций (поля country, card_type, cvv, threshold).
      - Сохранения нового отчёта (ReportsForm).

    Затем возвращает данные (включая сгенерированные) в шаблон test.html.
    """
    # Параметры, которые мы передадим в шаблон
    predict = None               # Строка-резюме от функции generate
    data_as_list = []            # Список словарей, чтобы не было проблем с DataFrame
    threshold_value = 0.5        # Значение по умолчанию

    # Создаём форму для отчёта (или обработаем POST)
    form = ReportsForm()

    if request.method == 'POST':
        # 1) Если нажали кнопку "Сгенерировать 100 транзакций"
        if 'generate' in request.POST:
            country = request.POST.get('country', '')
            card_type = request.POST.get('card_type', '')
            cvv = request.POST.get('cvv', '')

            # Пробуем считать threshold
            try:
                threshold_value = float(request.POST.get('threshold', 0.5))
            except ValueError:
                threshold_value = 0.5

            # Вызываем нашу функцию generate (которая возвращает (predict_str, numeric_data))
            predict, df = generate(
                country=country,
                card_type=card_type,
                cvv=cvv,
                threshold=threshold_value,
                count=100  # или другое число, если нужно
            )

            # Преобразуем DataFrame в список словарей, чтобы в шаблоне {% if generated_data %} работало
            data_as_list = df.to_dict(orient='records')

        # 2) Если нажали кнопку "Сохранить отчёт"
        if 'add_report' in request.POST:
            form = ReportsForm(request.POST)
            if form.is_valid():
                report = form.save(commit=False)
                report.user = request.user
                report.save()
                messages.success(request, "Отчёт успешно добавлен!")
                # Перенаправляем (может быть на test или на другую страницу)
                return redirect('users:test')
            else:
                messages.error(request, "Произошла ошибка при сохранении отчёта.")
    else:
        # Если GET-запрос, просто создаём форму
        form = ReportsForm()

    # ===== Пример фильтрации отчётов (если нужно) =====
    user_reports = Reports.objects.filter(user=request.user)
    report_name = request.GET.get('report_name', '').strip()
    report_date = request.GET.get('report_date', '').strip()
    if report_name:
        user_reports = user_reports.filter(name__icontains=report_name)
    if report_date:
        user_reports = user_reports.filter(date=report_date)

    # ===== Пример фильтрации транзакций (если нужно) =====
    user_transactions = Transactions.objects.all()
    transaction_bin = request.GET.get('transaction_bin', '').strip()
    transaction_day = request.GET.get('transaction_day', '').strip()
    transaction_month = request.GET.get('transaction_month', '').strip()

    if transaction_bin:
        user_transactions = user_transactions.filter(bin__icontains=transaction_bin)
    if transaction_day:
        user_transactions = user_transactions.filter(Day=transaction_day)
    if transaction_month:
        user_transactions = user_transactions.filter(Month=transaction_month)

    # Подготавливаем контекст и рендерим в test.html
    context = {
        'form': form,
        'user_reports': user_reports,
        'user_transactions': user_transactions,

        # Данные, связанные с генерацией
        'predict': predict,
        'generated_data': data_as_list,    # список словарей (не DataFrame)
        'threshold': threshold_value
    }
    return render(request, 'users/test.html', context)




def tex(request):
    # Обработка POST-запроса для добавления отчёта
    if request.method == 'POST' and 'add_report' in request.POST:
        form = ReportsForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            report.save()
            log_notification(request.user, "Вы добавили новый отчёт.")
            messages.success(request, "Отчёт успешно добавлен!")
            return redirect('users:test')
        else:
            messages.error(request, "Произошла ошибка при сохранении отчёта.")
    else:
        form = ReportsForm()

    # Фильтрация отчётов
    user_reports = Reports.objects.filter(user=request.user)
    report_name = request.GET.get('report_name', '').strip()
    report_date = request.GET.get('report_date', '').strip()
    if report_name:
        user_reports = user_reports.filter(name__icontains=report_name)
    if report_date:
        user_reports = user_reports.filter(date=report_date)

    # Фильтрация транзакций
    user_transactions = Transactions.objects.all()
    transaction_bin = request.GET.get('transaction_bin', '').strip()
    transaction_day = request.GET.get('transaction_day', '').strip()
    transaction_month = request.GET.get('transaction_month', '').strip()
    if transaction_bin:
        user_transactions = user_transactions.filter(bin__icontains=transaction_bin)
    if transaction_day:
        user_transactions = user_transactions.filter(Day=transaction_day)
    if transaction_month:
        user_transactions = user_transactions.filter(Month=transaction_month)

    # Проверка роли пользователя
    show_elements = False
    if request.user.role:
        show_elements = (request.user.role.name == "one")

    context = {
        'form': form,
        'user_reports': user_reports,
        'user_transactions': user_transactions,
        'show_elements': show_elements,
    }
    return render(request, 'users/tex.html', context)


# views.py
def test_one(request):
    # Если POST с добавлением отчёта
    if request.method == 'POST' and 'add_report' in request.POST:
        form = ReportsForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            report.save()
            log_notification(request.user, "Вы добавили новый отчёт.")
            messages.success(request, "Отчёт успешно добавлен!")
            return redirect('users:test')
        else:
            messages.error(request, "Произошла ошибка при сохранении отчёта.")
    else:
        form = ReportsForm()

    # Фильтрация отчётов
    user_reports = Reports.objects.filter(user=request.user)
    report_name = request.GET.get('report_name', '').strip()
    report_date = request.GET.get('report_date', '').strip()
    if report_name:
        user_reports = user_reports.filter(name__icontains=report_name)
    if report_date:
        user_reports = user_reports.filter(date=report_date)

    # Фильтрация транзакций
    user_transactions = Transactions.objects.all()
    transaction_bin = request.GET.get('transaction_bin', '').strip()
    transaction_day = request.GET.get('transaction_day', '').strip()
    transaction_month = request.GET.get('transaction_month', '').strip()
    if transaction_bin:
        user_transactions = user_transactions.filter(bin__icontains=transaction_bin)
    if transaction_day:
        user_transactions = user_transactions.filter(Day=transaction_day)
    if transaction_month:
        user_transactions = user_transactions.filter(Month=transaction_month)

    # Проверка роли пользователя (сначала проверяем, что role не None)
    show_elements = False
    if getattr(request.user, 'role', None):
        show_elements = (request.user.role.name == "one")

    context = {
        'form': form,
        'user_reports': user_reports,
        'user_transactions': user_transactions,
        'show_elements': show_elements,
    }

    return render(request, 'users/tex_one.html', context)




def tex_one(request):
    # Обработка POST-запроса для добавления отчёта
    if request.method == 'POST' and 'add_report' in request.POST:
        form = ReportsForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.user = request.user
            report.save()
            log_notification(request.user, "Вы добавили новый отчёт.")
            messages.success(request, "Отчёт успешно добавлен!")
            return redirect('users:test')
        else:
            messages.error(request, "Произошла ошибка при сохранении отчёта.")
    else:
        form = ReportsForm()

    # Фильтрация отчётов
    user_reports = Reports.objects.filter(user=request.user)
    report_name = request.GET.get('report_name', '').strip()
    report_date = request.GET.get('report_date', '').strip()
    if report_name:
        user_reports = user_reports.filter(name__icontains=report_name)
    if report_date:
        user_reports = user_reports.filter(date=report_date)

    # Фильтрация транзакций (result = "Не известно")
    user_transactions = Transactions.objects.filter(result="Не известно")
    transaction_bin = request.GET.get('transaction_bin', '').strip()
    transaction_day = request.GET.get('transaction_day', '').strip()
    transaction_month = request.GET.get('transaction_month', '').strip()
    if transaction_bin:
        user_transactions = user_transactions.filter(bin__icontains=transaction_bin)
    if transaction_day:
        user_transactions = user_transactions.filter(Day=transaction_day)
    if transaction_month:
        user_transactions = user_transactions.filter(Month=transaction_month)

    # Проверка роли пользователя
    show_elements = False
    if request.user.role:
        show_elements = (request.user.role.name == "one")

    context = {
        'form': form,
        'user_reports': user_reports,
        'user_transactions': user_transactions,
        'show_elements': show_elements,
    }
    return render(request, 'users/tex_two.html', context)


def update_transaction(request, transaction_id):
    """Обработчик изменения поля result для конкретной транзакции."""
    if request.method == "POST":
        transaction = get_object_or_404(Transactions, pk=transaction_id)
        new_result = request.POST.get('result')
        transaction.result = new_result
        transaction.save()
        log_notification(request.user, f"Вы обновили транзакцию с ID {transaction_id}.")
        messages.success(request, f"Транзакция {transaction_id} успешно обновлена!")
    return redirect('users:tex_one')


@login_required
def reports_view(request):
    """
    Страница «Отчёты», где можно:
    1) Отфильтровать транзакции.
    2) Посмотреть простую аналитику.
    3) Сформировать Excel-файл.
    """
    # Фильтрация транзакций
    transactions = Transactions.objects.all()

    bin_value = request.GET.get('bin_value', '').strip()
    month_value = request.GET.get('month_value', '').strip()

    if bin_value:
        transactions = transactions.filter(bin__icontains=bin_value)
    if month_value:
        transactions = transactions.filter(Month=month_value)

    total_count = transactions.count()
    sum_amount = 0
    for t in transactions:
        try:
            sum_amount += float(t.amount)
        except ValueError:
            pass

    context = {
        'transactions': transactions,
        'total_count': total_count,
        'sum_amount': sum_amount,
    }
    return render(request, 'users/reports.html', context)


@login_required
def download_excel_report(request):
    """
    Генерация и скачивание Excel-файла с аналитикой.
    """
    transactions = Transactions.objects.all()
    bin_value = request.GET.get('bin_value', '').strip()
    month_value = request.GET.get('month_value', '').strip()

    if bin_value:
        transactions = transactions.filter(bin__icontains=bin_value)
    if month_value:
        transactions = transactions.filter(Month=month_value)

    data = []
    for t in transactions:
        data.append({
            'bin': t.bin,
            'amount': t.amount,
            'shoppercountrycode': t.shoppercountrycode,
            'Day': t.Day,
            'Month': t.Month,
            'result': t.result,
        })
    df = pd.DataFrame(data)

    def to_float_safe(val):
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0

    df['amount'] = df['amount'].apply(to_float_safe)

    report_type = request.GET.get('report_type', 'by_country')

    wb = openpyxl.Workbook()

    if report_type == 'by_country':
        fraud_df = df[df['result'] == 'Мошенническая']
        report_df = fraud_df.groupby('shoppercountrycode').agg(
            fraud_count=('bin', 'count'),
            total_fraud_amount=('amount', 'sum')
        ).reset_index().sort_values('fraud_count', ascending=False)

        ws = wb.active
        ws.title = "Страны"
        headers = ["Страна", "Кол-во мошеннических транзакций", "Сумма мошеннических транзакций"]
        for col_num, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=h)
        for row_idx, row_data in enumerate(report_df.values, start=2):
            ws.cell(row=row_idx, column=1, value=row_data[0])
            ws.cell(row=row_idx, column=2, value=row_data[1])
            ws.cell(row=row_idx, column=3, value=row_data[2])
        filename = "report_by_country.xlsx"

    elif report_type == 'by_month':
        fraud_df = df[df['result'] == 'Не мошенническая']
        report_df = fraud_df.groupby('Month').agg(
            fraud_count=('bin', 'count'),
            total_fraud_amount=('amount', 'sum')
        ).reset_index().sort_values('Month')

        ws = wb.active
        ws.title = "По месяцам"
        headers = ["Месяц", "Кол-во мошеннических транзакций", "Сумма мошеннических транзакций"]
        for col_num, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=h)
        for row_idx, row_data in enumerate(report_df.values, start=2):
            ws.cell(row=row_idx, column=1, value=row_data[0])
            ws.cell(row=row_idx, column=2, value=row_data[1])
            ws.cell(row=row_idx, column=3, value=row_data[2])
        filename = "report_by_month.xlsx"

    else:
        ws = wb.active
        ws.title = "Общий отчёт"
        headers = df.columns.tolist()
        for col_num, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=h)
        for row_idx, row_data in enumerate(df.values, start=2):
            for col_num, cell_val in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_num, value=cell_val)
        filename = "report_general.xlsx"

    # Записываем уведомление о скачивании отчёта
    if request.user.is_authenticated:
        log_notification(request.user, "Вы скачали Excel-отчёт.")

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
